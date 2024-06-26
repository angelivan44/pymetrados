from os import X_OK
from numpy import count_nonzero
import openpyxl as xl
from openpyxl import load_workbook
from win32com.client import Dispatch
import re


def change_function(arg):
        return arg.value


def fun(arg, data):
    if (arg in data):
        return True
    else:
        return False

def sort_func(arg):
    return arg['id']

def get_uniques_values(data):
    count_data = list(map(change_function, data))
    new_set = set(count_data)
    if None in new_set:
        new_set.remove(None)
    new_array = []
    for i in new_set:
        if i != "Id":
            new_array.append({"id": i, "size": len(list(filter(lambda x: x == i, count_data)))})
    sorted_array = sorted(new_array, key=lambda x: x['id'])
    
    return sorted_array

def metados_py(planilla,db,export):
    error_message = None
    xl = None  # Inicializar la variable xl_app para poder referenciarla después

    #get paths
    try:
        planilla_path = planilla
        db_path = db
        result_path = export+'/Metrado.xlsx'

        #setup openpyxl for win32com

        xl = Dispatch("Excel.Application")

        #open workboos
        bd_workbook = xl.Workbooks.Open(Filename=db_path)
        result_workbook = xl.Workbooks.Add()
        datos_workbook = xl.Workbooks.Open(Filename=planilla_path)

        #open worksheets for xl and win32com

        result_worksheet = result_workbook.Worksheets(1)
        planilla_sheet = bd_workbook.Worksheets("P_0")
        metrado_sheet = bd_workbook.Worksheets("M_0")
        planilla_workbook = load_workbook(filename =planilla_path)
        data_base_workbook = load_workbook(filename=db )
        #get worksheets
        data = planilla_workbook["DB"]
        data_headers = planilla_workbook["PLANILLA"]
        #get data for planilla worksheets
        range_ids = data['A':'A']
        data_general = data_base_workbook["General"]
        configuracion = data_base_workbook["Configuracion"]
        range_dates = data_general["A2:E"+str(data_general.max_row)]
        dic_data = {}
        # Intenta ejecutar el bloque de código que podría fallar
            # Intenta ejecutar el bloque de código que podría fallar
        range_dates_general = list(map(lambda x: {"id": x[0].value, "name": x[1].value, "code": x[2].value, "system": x[3].value, "location": x[4].value}, list(range_dates)))
        dic_data = {data['id']: data for data in range_dates_general}
        # Captura cualquier excepción y almacena el mensaje de error en la variable
   
       

        name_location = configuracion['B2'].value
        id_location = configuracion['B3'].value
        system_location = configuracion['B4'].value
        ubication_location = configuracion['B5'].value
        first_row_planilla = configuracion['B6'].value
        match = re.match(r"([a-zA-Z]+)([0-9]+)", first_row_planilla)
        first_number_row_planilla = match.group(2)
        length_elements = configuracion['B7'].value
        first_row_montaje = configuracion['B8'].value
        first_value_montaje = configuracion['B9'].value
        range_rows_planilla = configuracion['B10'].value
        range_headers = data_headers['A1':range_rows_planilla][0]
        projects_ids = get_uniques_values(range_ids)
        count_headers = len(get_uniques_values(range_headers))
        acum = 0
        for item in projects_ids:
            try:
                data_project = dic_data[item['id']]
            except Exception as e:
                error_message = f"No se encontró el ID '{item['id']}' correspondiente a la localidad en la base de datos general. Por favor, verifica que el ID esté presente en la base de datos y que no haya errores de tipografía o formato. Asegúrate de que todas las localidades tienen un ID único y coincidente en todas las hojas de trabajo relacionadas."
                return error_message

            planilla_sheet.Copy(Before= result_worksheet)
            new_planilla = result_workbook.Worksheets("P_0")
            new_planilla.Name = f"P_{item['id']}"
            new_planilla.Range(name_location).Value = data_project['name']
            new_planilla.Range(id_location).Value = data_project['code']
            new_planilla.Range(system_location).Value = data_project['system']
            new_planilla.Range(ubication_location).Value = data_project['location']
            new_planilla.Range(first_row_planilla).Offset(2).EntireRow.GetResize(item['size']).Insert(Shift=-4121)
            new_planilla.Range(length_elements).Value = item['size']
            datos_worksheet = datos_workbook.Worksheets("PLANILLA")
            initial_value = 2 + acum
            acum = initial_value + item['size'] - 2
            last_value = initial_value + item['size'] - 1
            datos_worksheet.Range(datos_worksheet.Cells(initial_value,1), datos_worksheet.Cells(last_value,count_headers)).Copy()
            new_planilla.Range(first_row_planilla).Offset(1,2).PasteSpecial(Paste=-4163)

            metrado_sheet.Copy(Before= result_worksheet)
            montaje_data = result_workbook.Worksheets("M_0")
            montaje_data.Name = f"M_{item['id']}"
            montaje_data.Range(f"A{first_row_montaje}").Value = data_project['id']
            montaje_data.Range(f"C{first_row_montaje}").Value = data_project['name']
            try:
                montaje_data.Range(montaje_data.Cells(1,9),montaje_data.Cells(count_headers,item['size']+9)).FormulaArray = "=TRANSPOSE(" +"P_"+str(item['id']) + "!R["+ str(int(first_number_row_planilla) - 1) +"]C[-7]:R[" + str(14+item['size']) + "]C["+ str(count_headers-6) + "])"
                montaje_data.Range(montaje_data.Cells(first_row_montaje,9), montaje_data.Cells(1000,9)).AutoFill(montaje_data.Range(montaje_data.Cells(first_row_montaje,9), montaje_data.Cells(1000,8+item['size'])), 0 )
            except Exception as e:
                error_message = "Error: La versión de Excel no esta en Ingles. Por favor, cambie la configuración regional de Excel a inglés"
                return error_message

        metrado_sheet.Copy(Before= result_workbook.Worksheets(1))
        resumen_planilla = result_workbook.Worksheets("M_0")
        resumen_planilla.Name = "Resumen"
        max_row = data_base_workbook["M_0"].max_row
        for idx, project in enumerate(range_dates_general):
            resumen_planilla.Range(f"I{int(first_row_montaje)+3}").Offset(1,1+idx).Value = project['name']
            resumen_planilla.Range(f"I{int(first_row_montaje)+3}").Offset(2,1+idx).Value = project['id']
            resumen_planilla.Range(f"I{int(first_row_montaje)+3}").Offset(3,1+idx).Formula  = f"=+M_{project['id']}!{first_value_montaje}"
        resumen_planilla.Range(f"I{int(first_row_montaje)+6}")
        resumen_planilla.Range(resumen_planilla.Cells(int(first_row_montaje)+5,9), resumen_planilla.Cells(int(first_row_montaje)+5,9+len(range_dates_general))).AutoFill(resumen_planilla.Range(resumen_planilla.Cells(int(first_row_montaje)+5,9), resumen_planilla.Cells(max_row,9+len(range_dates_general))), 0 )
        resumen_planilla.Range(resumen_planilla.Cells(int(first_row_montaje)+5,7), resumen_planilla.Cells(max_row,7)).Copy()
        resumen_planilla.Range(resumen_planilla.Cells(int(first_row_montaje)+5,9), resumen_planilla.Cells(max_row,9+len(range_dates_general))).PasteSpecial(Paste=-4122)
        resumen_planilla.Range(resumen_planilla.Cells(int(first_row_montaje)+2,9), resumen_planilla.Cells(int(first_row_montaje)+4,9)).Copy()
        resumen_planilla.Range(resumen_planilla.Cells(int(first_row_montaje)+2,9), resumen_planilla.Cells(int(first_row_montaje)+4,8 + len(range_dates_general))).PasteSpecial(Paste=-4122)
        resumen_planilla.Range(f"A1:A{int(first_row_montaje)-2}").EntireRow.Delete()
        xl.DisplayAlerts = False
        dummy_sheet = result_workbook.Worksheets("Sheet1")
        dummy_sheet.Delete()
        bd_workbook.Close(SaveChanges=False)
        datos_workbook.Close(SaveChanges=False)
        result_workbook.SaveAs(result_path.replace('/', '\\'), ConflictResolution=2)
        result_workbook.Close()
        xl.Quit()
        return True
    except Exception as e:
        error_message = str(e)
        return error_message
    finally:
        # El bloque finally se ejecuta siempre, haya ocurrido un error o no.
        if xl:
            # Esto cerrará Excel si aún está abierto.
            xl.DisplayAlerts = False
            xl.Quit()